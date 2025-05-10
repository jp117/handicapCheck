import sys
import datetime
import requests # type: ignore
import csv
import openpyxl # type: ignore
import os
from dotenv import load_dotenv
from pathlib import Path
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
import pickle
import time
from email.mime.text import MIMEText
import base64
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

# Load environment variables
load_dotenv()

# Parse date with flexible year format
date_str = sys.argv[1]
try:
    # Try 4-digit year format first
    date = datetime.datetime.strptime(date_str, '%m-%d-%Y').date()
except ValueError:
    # If that fails, try 2-digit year format
    date = datetime.datetime.strptime(date_str, '%m-%d-%y').date()

# Get API key from environment variable
api_key = os.getenv('MTECH_API_KEY')
mtechAPIUrl = f'https://www.clubmtech.com/cmtapi/teetimes/?apikey={api_key}&TheDate={date.month}-{date.day}-{date.year}'

# Update SCOPES to include Google Sheets
SCOPES = [
    'https://www.googleapis.com/auth/gmail.readonly',
    'https://www.googleapis.com/auth/gmail.send',
    'https://www.googleapis.com/auth/spreadsheets'
]

# Add at the top with other global variables
cached_sheets = {}

def get_google_creds():
    """Get and refresh Google credentials."""
    creds = None
    if os.path.exists('token.json'):
        with open('token.json', 'rb') as token:
            creds = pickle.load(token)
    
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open('token.json', 'wb') as token:
            pickle.dump(creds, token)
    return creds

def update_google_sheet(sheet_id, sheet_name, golfer_list):
    """Update Google Sheet with golfer data."""
    global cached_sheets
    creds = get_google_creds()
    service = build('sheets', 'v4', credentials=creds)
    today_date = datetime.date.strftime(date, "%m-%d-%y")
    
    try:
        # Use cached data if available
        if sheet_name in cached_sheets:
            existing_data = cached_sheets[sheet_name]
        else:
            # Get existing data
            result = service.spreadsheets().values().get(
                spreadsheetId=sheet_id,
                range=f'{sheet_name}!A:C'
            ).execute()
            existing_data = result.get('values', [])
            cached_sheets[sheet_name] = existing_data
            
        if not existing_data:
            existing_data = [['Name', 'Count', 'Dates']]
        
        # Convert to dictionary for easier lookup
        data_dict = {row[0]: {'count': int(row[1]), 'dates': row[2]} 
                    for row in existing_data[1:] if len(row) >= 3}
        
        # Update data
        for golfer_name in golfer_list:
            golfer_name = golfer_name.strip() if isinstance(golfer_name, str) else str(golfer_name)
            
            if golfer_name in data_dict:
                # Update existing golfer
                data_dict[golfer_name]['count'] += 1
                current_dates = data_dict[golfer_name]['dates']
                data_dict[golfer_name]['dates'] = f"{current_dates}, {today_date}"
            else:
                # Add new golfer
                data_dict[golfer_name] = {'count': 1, 'dates': today_date}
        
        # Convert back to list format and sort by count
        updated_data = [[name, str(data['count']), data['dates']] 
                       for name, data in data_dict.items()]
        updated_data.sort(key=lambda x: int(x[1]), reverse=True)
        
        # Add header
        final_data = [['Name', 'Count', 'Dates']] + updated_data
        
        # Clear existing data
        service.spreadsheets().values().clear(
            spreadsheetId=sheet_id,
            range=f'{sheet_name}!A:C'
        ).execute()
        
        # Update sheet
        service.spreadsheets().values().update(
            spreadsheetId=sheet_id,
            range=f'{sheet_name}!A1',
            valueInputOption='USER_ENTERED',
            body={'values': final_data}
        ).execute()
        
    except Exception as e:
        print(f"Error updating Google Sheet: {e}")

def test_gmail_connection():
    """Test Gmail API connection and list recent emails."""
    creds = None
    # The file token.json stores the user's access and refresh tokens
    if os.path.exists('token.json'):
        with open('token.json', 'rb') as token:
            creds = pickle.load(token)
    
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.json', 'wb') as token:
            pickle.dump(creds, token)

    try:
        # Create Gmail API service
        service = build('gmail', 'v1', credentials=creds)
        
        # Get list of messages
        results = service.users().messages().list(userId='nhcchandicapcheck@gmail.com', maxResults=5).execute()
        messages = results.get('messages', [])

        if not messages:
            return True
        
        return True
    except Exception as e:
        print(f'An error occurred: {e}')
        return False

def removeAfterCharacter(text, char):
    index = text.find(char)
    if index != -1:
        return text[:index]
    return text


def getMTechData():
    with requests.Session() as s:
        download = s.get(mtechAPIUrl)

        decodedContet = download.content.decode('utf-8')

        teeSheet = csv.reader(decodedContet.splitlines(), delimiter=",")
        teeSheet = list(teeSheet)
        teeSheet = teeSheet[1:]
        
        return teeSheet

def get_email_attachment(search_date):
    """Get XLSX attachment from GHIN email for the specified date."""
    creds = get_google_creds()
    service = build('gmail', 'v1', credentials=creds)
    
    # Format date for email search (next day after golf date)
    email_date = search_date + datetime.timedelta(days=1)
    search_query = f'from:reporting@ghin.com subject:"Auto-Generated Scheduled Report - Played / Posted Report (Player Rounds)" after:{email_date.strftime("%Y/%m/%d")} before:{(email_date + datetime.timedelta(days=1)).strftime("%Y/%m/%d")}'
    
    try:
        # Search for the email
        results = service.users().messages().list(userId='me', q=search_query).execute()
        messages = results.get('messages', [])

        if not messages:
            raise Exception(f"No email found from reporting@ghin.com for date {email_date.strftime('%m-%d-%y')}")

        # Get the first matching email
        msg = service.users().messages().get(userId='me', id=messages[0]['id']).execute()

        # Get the attachment
        if 'parts' not in msg['payload']:
            raise Exception("Email has no attachments")

        for part in msg['payload']['parts']:
            if part.get('filename', '').endswith('.xlsx'):
                attachment_id = part['body']['attachmentId']
                attachment = service.users().messages().attachments().get(
                    userId='me', 
                    messageId=messages[0]['id'], 
                    id=attachment_id
                ).execute()
                
                # Decode attachment data
                file_data = attachment['data']
                file_data = file_data.replace('-', '+').replace('_', '/')
                
                # Create a temporary file to store the Excel data
                import tempfile
                
                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                temp_file.write(base64.urlsafe_b64decode(file_data))
                temp_file.close()
                
                return temp_file.name
                
        raise Exception("No XLSX attachment found in email")
        
    except Exception as e:
        print(f"Error getting email attachment: {e}")
        raise

def getUSGAData():
    """Get USGA data once and cache it."""
    if not hasattr(getUSGAData, 'cached_data'):
        try:
            # Get the XLSX file from email
            excel_file = get_email_attachment(date)
            
            # Read the Excel file
            postData = openpyxl.load_workbook(excel_file)
            sheet = postData.active
            golfers = []
            for row in sheet:
                golfers.append([cell.value for cell in row])
            
            # Clean up temporary file
            os.unlink(excel_file)
            
            # Cache the data
            getUSGAData.cached_data = golfers[1:]  # Skip header row
            
        except Exception as e:
            print(f"Error reading USGA data: {e}")
            raise
            
    return getUSGAData.cached_data

def checkPosting(golfer):
    """Check if a golfer has posted their score."""
    try:
        usga_data = getUSGAData()  # This will now use cached data after first call
        golfer_id = int(golfer)
        
        for usga in usga_data:
            if int(usga[0]) == golfer_id:
                return True
        return False
    except Exception as e:
        print(f"Error checking posting: {e}")
        return False

def normalize_name(name):
    """Normalize name by removing extra spaces and special characters"""
    return ' '.join(name.strip().lower().split())

def cache_sheet_data():
    """Cache all sheet data at startup to avoid rate limits"""
    creds = get_google_creds()
    service = build('sheets', 'v4', credentials=creds)
    
    # Cache all needed data
    try:
        # Get excluded dates
        result = service.spreadsheets().values().get(
            spreadsheetId=os.getenv('GOOGLE_SHEET_ID'),
            range='ExcludedDates!A:C'
        ).execute()
        cache_sheet_data.excluded_dates = result.get('values', [])
        
        # Get roster data
        result = service.spreadsheets().values().get(
            spreadsheetId=os.getenv('ROSTER_SHEET_ID'),
            range='Sheet1!A:E'
        ).execute()
        cache_sheet_data.roster = result.get('values', [])
        
        return True
    except Exception as e:
        print(f"Error caching sheet data: {e}")
        return False

def get_sheet(sheet_name):
    """Get a specific sheet from the Google Spreadsheet with caching."""
    global cached_sheets
    
    # Return cached data if available
    if sheet_name in cached_sheets:
        return cached_sheets[sheet_name]
        
    try:
        creds = get_google_creds()
        service = build('sheets', 'v4', credentials=creds)
        
        # Get the sheet data
        result = service.spreadsheets().values().get(
            spreadsheetId=os.getenv('GOOGLE_SHEET_ID'),
            range=f'{sheet_name}!A:C'
        ).execute()
        
        # Cache the result
        cached_sheets[sheet_name] = result.get('values', [])
        return cached_sheets[sheet_name]
    except Exception as e:
        return None

def get_excluded_times(date_str):
    excluded_times = []
    try:
        # Get the ExcludedDates sheet
        values = get_sheet('ExcludedDates')
        if not values:
            return excluded_times

        # Skip header row
        for row in values[1:]:
            if len(row) >= 1 and row[0] == date_str:
                # If there are no times specified, exclude the entire day
                if len(row) == 1 or (len(row) > 1 and not row[1].strip() and (len(row) <= 2 or not row[2].strip())):
                    excluded_times.append((None, None))
                else:
                    start_time = row[1] if len(row) > 1 and row[1] else None
                    end_time = row[2] if len(row) > 2 and row[2] else None
                    excluded_times.append((start_time, end_time))
    except Exception as e:
        return excluded_times
    return excluded_times

def check_roster(golfer_name):
    """Check if golfer exists in roster and has a GHIN number"""
    if not hasattr(cache_sheet_data, 'roster'):
        return False, False
        
    normalized_golfer = normalize_name(golfer_name)
    
    for row in cache_sheet_data.roster[1:]:  # Skip header
        if not row[0]:  # Skip empty names
            continue
            
        if normalize_name(row[0]) == normalized_golfer:
            has_ghin = len(row) > 1 and row[1] and row[1].strip()
            return True, has_ghin
            
    return False, False

def parse_mtech_time(time_str):
    try:
        parsed_time = datetime.datetime.strptime(time_str, '%I:%M %p').time()
        return parsed_time
    except Exception as e:
        return None

def is_time_excluded(tee_time):
    try:
        parsed_time = parse_mtech_time(tee_time)
        if not parsed_time:
            return False

        date_str = get_current_date()
        excluded_times = get_excluded_times(date_str)

        for start_time_str, end_time_str in excluded_times:
            # If both times are None, it's a full day exclusion
            if start_time_str is None and end_time_str is None:
                return True

            if start_time_str:
                start_time = datetime.datetime.strptime(start_time_str, '%H:%M').time()
            else:
                start_time = None

            if end_time_str:
                end_time = datetime.datetime.strptime(end_time_str, '%H:%M').time()
            else:
                end_time = None

            # Check if the tee time falls within the exclusion period
            if (start_time is None or parsed_time >= start_time) and \
               (end_time is None or parsed_time <= end_time):
                return True

        return False
    except Exception as e:
        return False

def get_current_date():
    """Get the current date in the format used in the ExcludedDates sheet"""
    return date.strftime("%m-%d-%y")  # Make sure we're using the right format

def get_roster_info(golfer_name):
    """Return (exists, has_ghin, email, gender, member_number) for a golfer."""
    if not hasattr(cache_sheet_data, 'roster'):
        return False, False, None, None, None

    normalized_golfer = normalize_name(golfer_name)
    for row in cache_sheet_data.roster[1:]:  # Skip header
        if not row[0]:
            continue
        if normalize_name(row[0]) == normalized_golfer:
            has_ghin = len(row) > 1 and row[1] and row[1].strip()
            email = row[2].strip() if len(row) > 2 and row[2] else None
            gender = row[3].strip().upper() if len(row) > 3 and row[3] else None
            member_number = row[4].strip() if len(row) > 4 and row[4] else None
            return True, has_ghin, email, gender, member_number
    return False, False, None, None, None

def build_no_post_email(men, women, date_str):
    """Create separate Excel files for men and women non-posting golfers."""
    files = []
    
    # Function to create a single Excel file
    def create_excel_file(golfers, gender):
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Add headers
        ws['A1'] = 'Last Name'
        ws['B1'] = 'MemberNo'
        ws['C1'] = 'Email'
        ws['D1'] = 'NO_POST_DATE'
        
        # Add data
        row = 2
        for name, email, member_number in golfers:
            ws[f'A{row}'] = name
            ws[f'B{row}'] = member_number or ''
            ws[f'C{row}'] = email or ''
            ws[f'D{row}'] = date_str
            row += 1
        
        # Save to a temporary file
        temp_file = f'reports/non_posters_{gender}_{date_str.replace("-", "_")}.xlsx'
        Path('reports').mkdir(parents=True, exist_ok=True)
        wb.save(temp_file)
        return temp_file
    
    # Create separate files for men and women
    if men:
        files.append(create_excel_file(men, 'men'))
    if women:
        files.append(create_excel_file(women, 'women'))
    
    return files

def send_email(subject, body, to_email, attachment_paths=None):
    """Send email with optional Excel attachments."""
    creds = get_google_creds()
    service = build('gmail', 'v1', credentials=creds)
    
    # Create message container
    message = MIMEMultipart()
    message['to'] = to_email
    message['from'] = "me"
    message['subject'] = subject
    
    # Add body text
    message.attach(MIMEText(body))
    
    # Add attachments if provided
    if attachment_paths:
        for attachment_path in attachment_paths:
            with open(attachment_path, 'rb') as f:
                attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                attachment.set_payload(f.read())
                encoders.encode_base64(attachment)
                attachment.add_header('Content-Disposition', 'attachment', 
                                    filename=os.path.basename(attachment_path))
                message.attach(attachment)
    
    # Encode and send
    raw = base64.urlsafe_b64encode(message.as_bytes()).decode()
    message_body = {'raw': raw}
    service.users().messages().send(userId="me", body=message_body).execute()
    
    # Clean up temporary files
    if attachment_paths:
        for path in attachment_paths:
            if os.path.exists(path):
                os.remove(path)

def get_roster_name_by_ghin(ghin_number):
    """Look up a golfer's name in the roster by their GHIN number."""
    if not hasattr(cache_sheet_data, 'roster'):
        return None
        
    ghin_number = str(ghin_number).strip()
    for row in cache_sheet_data.roster[1:]:  # Skip header
        if len(row) > 1 and row[1] and str(row[1]).strip() == ghin_number:
            return row[0]  # Return name from column A
    return None

if __name__ == "__main__":
    # Cache sheet data at startup
    print("Caching sheet data...")
    if not cache_sheet_data():
        print("Failed to cache sheet data. Exiting.")
        sys.exit(1)
    
    # Test Gmail connection first
    print("Testing Gmail connection...")
    test_gmail_connection()
    
    noGHIN = []
    noPost = []
    tee_data = getMTechData()
    
    print(f"Checking golf rounds for {date.strftime('%m-%d-%y')}")
    print(f"Looking for USGA report email from {(date + datetime.timedelta(days=1)).strftime('%m-%d-%y')}")
    
    # Collect all golfer[1] values to check for duplicates
    all_golfer_values = [g[1] for g in tee_data if len(g) > 1]
    
    men_no_post = []
    women_no_post = []
    posted_golfers = []

    checked_golfers = set()

    for golfer in tee_data:
        # Use GHIN if available, otherwise fallback to name
        unique_id = golfer[3] if golfer[3] else removeAfterCharacter(golfer[2], '-')
        if unique_id in checked_golfers:
            continue
        checked_golfers.add(unique_id)

        if len(golfer) > 1 and all_golfer_values.count(golfer[1]) > 1:
            if not is_time_excluded(golfer[1]):
                # Get the canonical name from roster if GHIN exists
                canonical_name = None
                if golfer[3]:  # If GHIN exists in MTech data
                    canonical_name = get_roster_name_by_ghin(golfer[3])
                if not canonical_name:  # Fallback to MTech name if no GHIN or not found in roster
                    canonical_name = removeAfterCharacter(golfer[2], '-')
                
                if golfer[3] != "":
                    if checkPosting(golfer[3]):
                        posted_golfers.append(canonical_name)
                    else:
                        noPost.append(canonical_name)
                        exists, has_ghin, email, gender, member_number = get_roster_info(canonical_name)
                        gender = (gender or '').strip().upper()
                        if gender == "M":
                            men_no_post.append((canonical_name, email, member_number))
                        elif gender == "F":
                            women_no_post.append((canonical_name, email, member_number))
                        else:
                            pass
                elif golfer[3] == "":
                    golfer_name = removeAfterCharacter(golfer[2], "-")
                    exists_in_roster, has_ghin, email, gender, member_number = get_roster_info(golfer_name)
                    if exists_in_roster and has_ghin:
                        for row in cache_sheet_data.roster[1:]:
                            if normalize_name(row[0]) == normalize_name(golfer_name):
                                roster_ghin = row[1].strip()
                                break
                        else:
                            roster_ghin = None
                        if roster_ghin and checkPosting(roster_ghin):
                            posted_golfers.append(golfer_name)
                        else:
                            noPost.append(golfer_name)
                            gender = (gender or '').strip().upper()
                            if gender == "M":
                                men_no_post.append((golfer_name, email, member_number))
                            elif gender == "F":
                                women_no_post.append((golfer_name, email, member_number))
                    else:
                        noGHIN.append(golfer_name)
    
    print('Handicap report done for ' + datetime.date.strftime(date, "%m-%d-%y"))
    
    # Make sure we have a reports directory
    Path('reports').mkdir(parents=True, exist_ok=True)
    
    # Replace the Excel report updates with Google Sheets updates
    SPREADSHEET_ID = os.getenv('GOOGLE_SHEET_ID')
    
    # Update both sheets
    update_google_sheet(SPREADSHEET_ID, 'NoPost', noPost)
    time.sleep(1)  # Add 1 second delay between updates
    update_google_sheet(SPREADSHEET_ID, 'NoGHIN', noGHIN)

    date_str = date.strftime("%m-%d-%y")
    excel_files = build_no_post_email(men_no_post, women_no_post, date_str)
    
    # Create appropriate message based on which files were created
    if len(excel_files) == 2:
        body = "Please find attached the Excel reports for both men and women non-posting golfers."
    elif len(excel_files) == 1:
        gender = "men" if men_no_post else "women"
        body = f"Please find attached the Excel report for {gender} non-posting golfers."
    else:
        body = "No non-posting golfers found for this date."
    
    send_email(
        subject=f"Non-Posters for {date_str}",
        body=body,
        to_email="John.Paradise117@gmail.com",
        attachment_paths=excel_files
    )
    print("Summary email sent with Excel attachment(s).")

    # Update PostPercentage sheet
    def update_post_percentage(sheet_id, posted_golfers, no_post_golfers):
        creds = get_google_creds()
        service = build('sheets', 'v4', credentials=creds)
        sheet_name = 'PostPercentage'
        try:
            result = service.spreadsheets().values().get(
                spreadsheetId=sheet_id,
                range=f'{sheet_name}!A:F'
            ).execute()
            existing_data = result.get('values', [])
            if not existing_data:
                existing_data = [['Name', 'Rounds Posted', 'Rounds Not Posted', 'Other', 'Pct All', 'Pct Played']]
            data_dict = {row[0]: row for row in existing_data[1:] if len(row) > 0}
            # Update posted
            for name in posted_golfers:
                if name in data_dict:
                    row = data_dict[name]
                    # Ensure row has at least 3 columns
                    while len(row) < 6:
                        row.append('')
                    # Increment posted count (col B)
                    row[1] = str(int(row[1]) + 1) if len(row) > 1 and row[1].isdigit() else '1'
                else:
                    data_dict[name] = [name, '1', '', '', '', '']
            # Update not posted
            for name in no_post_golfers:
                if name in data_dict:
                    row = data_dict[name]
                    # Ensure row has at least 3 columns
                    while len(row) < 6:
                        row.append('')
                    # Increment not posted count (col C)
                    row[2] = str(int(row[2]) + 1) if len(row) > 2 and row[2].isdigit() else '1'
                else:
                    data_dict[name] = [name, '', '1', '', '', '']
            # Prepare final data
            final_data = [existing_data[0]]
            for row in data_dict.values():
                # Ensure row has at least 6 columns
                while len(row) < 6:
                    row.append('')
                # Row number in sheet is index+2 (header is row 1)
                row_idx = len(final_data) + 1
                row[4] = f"=B{row_idx}/(B{row_idx}+C{row_idx}+D{row_idx})"
                row[5] = f"=B{row_idx}/(B{row_idx}+C{row_idx})"
                final_data.append(row)
            # Update sheet
            service.spreadsheets().values().update(
                spreadsheetId=sheet_id,
                range=f'{sheet_name}!A1',
                valueInputOption='USER_ENTERED',
                body={'values': final_data}
            ).execute()
        except Exception as e:
            print(f"Error updating {sheet_name}: {e}")
    update_post_percentage(SPREADSHEET_ID, posted_golfers, noPost)